import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

data1 = {'A': [1111, 2222, 3333]}
df1 = pd.DataFrame(data1)
df1.to_excel('file1.xlsx', index=False)

data2 = {'B': [1111, 2222, 3333]}
df2 = pd.DataFrame(data2)
df2.to_excel('file2.xlsx', index=False)

data3 = {'C': [1111, 2222, 3333]}
df3 = pd.DataFrame(data3)
df3.to_excel('file3.xlsx', index=False)

file_paths = ['file1.xlsx', 'file2.xlsx', 'file3.xlsx']
dfs = [pd.read_excel(file) for file in file_paths]

result_df = pd.concat(dfs, axis=1)

result_df = result_df.sort_values(by=list(result_df.columns), ascending=False)

output_file_path = 'output_file.xlsx'
result_df.to_excel(output_file_path, index=False, header=False)

workbook = Workbook()
worksheet = workbook.active

font = Font(bold=True)
for row in worksheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.font = font

border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for row in worksheet.iter_rows():
    for cell in row:
        cell.border = border

for row_index, row in enumerate(result_df.values, start=1):
    for col_index, value in enumerate(row, start=1):
        worksheet.cell(row=row_index, column=col_index, value=value)

workbook.save(output_file_path)
